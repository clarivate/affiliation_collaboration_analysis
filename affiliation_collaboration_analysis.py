"""Analyze author affiliations in Web of Science Expanded API records.

The script identifies authors affiliated with a target organization and writes
record-level collaboration details to an Excel workbook. Results are saved after
each retrieved page so partial progress is preserved if a later request fails.
"""

import argparse
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import wosesrclient_robust
from wosesrclient_robust import InvalidWoSQueryError, WoSAuthenticationError


# Configuration

DEFAULT_QUERY = "OG=(Trinity College)"
DEFAULT_AFFILIATION = "Trinity College"
OUTPUT_PREFIX = "TrinColl"

COUNT = 50

AFFILIATION_WARNING_THRESHOLD = 500
MAIN_QUERY_WARNING_THRESHOLD = 25000

# Page sizes used when a response is too large or otherwise fails.
ADAPTIVE_STEP_UP_AFTER_PAGES = 5
PAGE_SIZE_FALLBACKS = [50, 25, 10, 5, 1]

# General helpers

def ensure_list(value: Any) -> List[Any]:
    """
    Normalize API fields that may appear as a dict when there is one item
    or as a list when there are multiple items.
    """
    if value is None:
        return []
    return value if isinstance(value, list) else [value]


def normalize(text: Optional[Any]) -> str:
    """
    Normalize strings for comparison.
    """
    if text is None:
        return ""
    return " ".join(str(text).strip().lower().split())


def safe_join(values: List[str], sep: str = " | ") -> str:
    """
    Join nonblank values for output.
    """
    return sep.join([v for v in values if v])


def sort_addr_no(value: str) -> Tuple[int, Any]:
    """Return a stable sort key for numeric and nonnumeric address IDs."""
    value_text = str(value)
    if value_text.isdigit():
        return 0, int(value_text)
    return 1, value_text


def make_default_output_filename(prefix: str = OUTPUT_PREFIX) -> Path:
    """Create a timestamped Excel output filename."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    return Path(f"{prefix}_{timestamp}.xlsx")


def is_internal_only_collab(total_authors: Any, matching_authors: int) -> str:
    """
    Returns YES when all authors on the record match the target affiliation.
    Otherwise returns NO.
    """
    try:
        return "YES" if int(total_authors) == int(matching_authors) else "NO"
    except (TypeError, ValueError):
        return "NO"


def make_og_query(affiliation: str) -> str:
    """
    Build a simple OG query for checking whether the target affiliation
    is likely valid in Web of Science.

    Example:
        University of Pittsburgh
    becomes:
        OG=(University of Pittsburgh)
    """
    clean_affiliation = str(affiliation).strip().replace('"', '\\"')
    return f"OG=({clean_affiliation})"


def confirm_continue_after_warning(message: str) -> bool:
    """
    Ask user to confirm continuation after a guardrail warning.
    """
    print()
    print("----- WARNING -----")
    print(message)
    response = input("Continue anyway? Type YES to continue: ").strip().upper()
    return response == "YES"


def build_page_size_fallbacks(preferred_count: int) -> List[int]:
    """
    Build a descending list of page sizes to try.

    Examples:
    - preferred_count=50 -> [50, 25, 10, 5, 1]
    - preferred_count=10 -> [10, 5, 1]
    - preferred_count=5  -> [5, 1]
    """

    candidates = [preferred_count] + PAGE_SIZE_FALLBACKS
    cleaned: List[int] = []

    for value in candidates:
        try:
            value = int(value)
        except (TypeError, ValueError):
            continue

        if value < 1:
            continue

        if value <= preferred_count and value not in cleaned:
            cleaned.append(value)

    cleaned = sorted(cleaned, reverse=True)

    if 1 not in cleaned:
        cleaned.append(1)

    return cleaned


def get_ordered_page_sizes(max_count: int) -> List[int]:
    """
    Return valid page sizes up to max_count, sorted ascending.

    Example:
    max_count=50 -> [1, 5, 10, 25, 50]
    max_count=10 -> [1, 5, 10]
    """
    sizes = sorted(set([1, 5, 10, 25, 50, max_count]))
    return [s for s in sizes if 1 <= s <= max_count]


def get_next_larger_page_size(current_count: int, max_count: int) -> int:
    """
    Step up one page-size level after enough successful pages.
    """
    sizes = get_ordered_page_sizes(max_count)

    for size in sizes:
        if size > current_count:
            return size

    return current_count

# Preflight checks

def get_records_found_for_query(apikey: str, usr_query: str) -> int:
    """
    Lightweight preflight query to get QueryResult.RecordsFound
    without retrieving the full result set.

    Uses wosesrclient_robust.get_response() so the preflight uses the same
    endpoint, headers, retries, and error handling as the main retrieval.
    """

    params = {
        "databaseId": "WOS",
        "usrQuery": usr_query
    }

    response = wosesrclient_robust.get_response(
        apikey=apikey,
        params=params,
        firstRecord=1,
        count=1
    )

    return int(
        response.get("QueryResult", {})
                .get("RecordsFound", 0)
    )


def run_affiliation_preflight_check(
    apikey: str,
    affiliation: str,
    threshold: int,
    auto_yes: bool
) -> bool:
    """
    Check whether the target affiliation appears to be a valid OG value.
    If the hit count is below the threshold, ask whether to continue.
    """

    affiliation_check_query = make_og_query(affiliation)

    print("Running affiliation preflight check...")
    print(f"Affiliation check query: {affiliation_check_query}")

    affiliation_hits = get_records_found_for_query(
        apikey=apikey,
        usr_query=affiliation_check_query
    )

    print(f"Affiliation check RecordsFound: {affiliation_hits}")

    if affiliation_hits < threshold:
        warning_message = (
            f'The target affiliation "{affiliation}" returned only '
            f'{affiliation_hits} records using:\n'
            f"  {affiliation_check_query}\n\n"
            "This may indicate a typo, variant institution name, or unexpected "
            "WoS organization form."
        )

        if auto_yes:
            print("WARNING acknowledged automatically because --yes was used.")
            return True

        return confirm_continue_after_warning(warning_message)

    return True


def run_main_query_count_preflight_check(
    apikey: str,
    usr_query: str,
    threshold: int,
    auto_yes: bool
) -> bool:
    """
    Check the main query size before retrieving all records.
    If the query is very large, ask whether to continue.
    """

    print("Running main query size preflight check...")

    main_query_hits = get_records_found_for_query(
        apikey=apikey,
        usr_query=usr_query
    )

    print(f"Main query RecordsFound: {main_query_hits}")

    if main_query_hits >= threshold:
        warning_message = (
            f"The main query returned {main_query_hits} records, which is at or "
            f"above the warning threshold of {threshold}.\n\n"
            "This may take a long time, make many API requests, and produce a "
            "large Excel file."
        )

        if auto_yes:
            print("WARNING acknowledged automatically because --yes was used.")
            return True

        return confirm_continue_after_warning(warning_message)

    return True


# API retrieval

def extract_records_found(response: Dict[str, Any]) -> int:
    """
    Extract QueryResult.RecordsFound from a WoS API response.
    """
    return int(
        response.get("QueryResult", {})
                .get("RecordsFound", 0)
    )


def extract_records_from_response(response: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Extract REC records from a WoS API response and normalize to a list.
    """
    try:
        recs = response["Data"]["Records"]["records"]["REC"]
    except Exception:
        return []

    if isinstance(recs, dict):
        return [recs]

    if isinstance(recs, list):
        return recs

    return []


def get_wos_page(
    apikey: str,
    params: Dict[str, Any],
    first_record: int,
    count: int
) -> Dict[str, Any]:
    """
    Retrieve one page of WoS results using the robust client.
    """
    return wosesrclient_robust.get_response(
        apikey=apikey,
        params=params,
        firstRecord=first_record,
        count=count
    )


def get_wos_page_with_auto_shrink(
    apikey: str,
    params: Dict[str, Any],
    first_record: int,
    preferred_count: int
) -> Tuple[Optional[Dict[str, Any]], int, Optional[str]]:
    """
    Retrieve one page of WoS results.

    If retrieval fails, automatically retry the same firstRecord with smaller
    page sizes. This helps with very large records, such as group-author
    collaboration records with huge author/address metadata.

    Returns:
    - response dict, or None if all attempts failed
    - count/page size that succeeded, or 1 if all attempts failed
    - error message if all attempts failed, otherwise None
    """

    fallback_counts = build_page_size_fallbacks(preferred_count)
    last_error = None

    for attempt_count in fallback_counts:
        try:
            if attempt_count != preferred_count:
                print(
                    f"Retrying firstRecord={first_record} with smaller count={attempt_count}"
                )

            response = get_wos_page(
                apikey=apikey,
                params=params,
                first_record=first_record,
                count=attempt_count
            )

            return response, attempt_count, None

        except MemoryError as e:
            last_error = (
                f"MemoryError at firstRecord={first_record}, "
                f"count={attempt_count}: {e}"
            )
            print(last_error)

        except Exception as e:
            last_error = (
                f"Retrieval error at firstRecord={first_record}, "
                f"count={attempt_count}: {e}"
            )
            print(last_error)

    return None, 1, last_error


# Record extraction

def extract_uid(rec: Dict[str, Any]) -> str:
    return rec.get("UID", "N/A")


def extract_item_title(rec: Dict[str, Any]) -> str:
    """
    Pull the item title from static_data.summary.titles.title.
    """
    titles = (
        rec.get("static_data", {})
           .get("summary", {})
           .get("titles", {})
           .get("title")
    )

    for title_obj in ensure_list(titles):
        if not isinstance(title_obj, dict):
            continue

        if title_obj.get("type") == "item":
            return str(title_obj.get("content", ""))

    return ""


def extract_local_count_by_coll_id(rec: Dict[str, Any], coll_id: str) -> Any:
    """
    Pull local_count from dynamic_data.citation_related.tc_list.silo_tc
    for the requested coll_id.

    Examples:
    - WOK = WOS Platform Cites
    - WOS = Core Collection Cites
    """
    silo_tc = (
        rec.get("dynamic_data", {})
           .get("citation_related", {})
           .get("tc_list", {})
           .get("silo_tc")
    )

    target_coll_id = str(coll_id).upper()

    for tc_obj in ensure_list(silo_tc):
        if not isinstance(tc_obj, dict):
            continue

        if str(tc_obj.get("coll_id", "")).upper() == target_coll_id:
            return tc_obj.get("local_count", 0)

    return 0


def extract_wos_platform_cites(rec: Dict[str, Any]) -> Any:
    """
    WOS Platform Cites = local_count where coll_id == 'WOK'.
    """
    return extract_local_count_by_coll_id(rec, "WOK")


def extract_core_collection_cites(rec: Dict[str, Any]) -> Any:
    """
    Core Collection Cites = local_count where coll_id == 'WOS'.
    """
    return extract_local_count_by_coll_id(rec, "WOS")


def get_total_authors(rec: Dict[str, Any]) -> Any:
    """
    Prefer summary.names.count for total authors.
    If missing, count author-role names in summary.names.name.
    """
    names_container = (
        rec.get("static_data", {})
           .get("summary", {})
           .get("names", {})
    )

    count = names_container.get("count")
    if count is not None:
        return count

    names = ensure_list(names_container.get("name"))
    author_count = 0

    for name_obj in names:
        if isinstance(name_obj, dict) and normalize(name_obj.get("role")) == "author":
            author_count += 1

    return author_count if author_count else "N/A"


def get_addresses_container(rec: Dict[str, Any]) -> Dict[str, Any]:
    """
    Return fullrecord_metadata.addresses if available.
    """
    addresses_container = (
        rec.get("static_data", {})
           .get("fullrecord_metadata", {})
           .get("addresses")
    )

    return addresses_container if isinstance(addresses_container, dict) else {}


def get_author_key(name_obj: Dict[str, Any]) -> Optional[str]:
    """
    Stable key for distinct author counting.

    Preference order:
    1. seq_no
    2. wos_standard
    3. full_name
    4. display_name
    """
    if not isinstance(name_obj, dict):
        return None

    key = (
        name_obj.get("seq_no")
        or name_obj.get("wos_standard")
        or name_obj.get("full_name")
        or name_obj.get("display_name")
    )

    if key is None:
        return None

    return str(key)


def get_author_display(name_obj: Dict[str, Any]) -> str:
    """
    Human-readable author label for output.
    """
    if not isinstance(name_obj, dict):
        return ""

    return str(
        name_obj.get("full_name")
        or name_obj.get("display_name")
        or name_obj.get("wos_standard")
        or name_obj.get("seq_no")
        or ""
    )


def extract_suborganizations(address_spec: Dict[str, Any]) -> List[str]:
    """
    Extract suborganization names from address_spec.suborganizations.

    Expected shapes can vary:
    1. {"suborganization": "Dept Occupat Therapy"}
    2. {"suborganization": ["Dept A", "Dept B"]}
    3. {"suborganization": {"content": "Dept A"}}
    4. {"suborganization": [{"content": "Dept A"}, {"content": "Dept B"}]}
    """
    if not isinstance(address_spec, dict):
        return []

    suborg_block = address_spec.get("suborganizations")
    if not isinstance(suborg_block, dict):
        return []

    raw_suborgs = ensure_list(suborg_block.get("suborganization"))

    suborgs: List[str] = []

    for item in raw_suborgs:
        if item is None:
            continue

        if isinstance(item, dict):
            value = item.get("content")
        else:
            value = item

        value = str(value).strip() if value is not None else ""

        if value:
            suborgs.append(value)

    return suborgs


def org_matches_target(orgs: List[Any], target_norm: str) -> bool:
    """
    Determine whether an address's organization list matches the target affiliation.

    First pass:
    - exact match on pref='Y' organization

    Second pass:
    - exact match on any organization
    """
    # Preferred organization exact match
    for org in orgs:
        if not isinstance(org, dict):
            continue

        pref = str(org.get("pref", "")).strip().upper()
        org_name = normalize(org.get("content"))

        if pref == "Y" and org_name == target_norm:
            return True

    # Fallback: any organization exact match
    for org in orgs:
        if not isinstance(org, dict):
            continue

        org_name = normalize(org.get("content"))

        if org_name == target_norm:
            return True

    return False


# Analysis

def analyze_affiliation(
    rec: Dict[str, Any],
    target_affiliation: str
) -> Dict[str, Any]:
    """
    Analyze one WoS record for target-affiliation collaboration.
    """

    target_norm = normalize(target_affiliation)

    total_authors = get_total_authors(rec)

    addresses_container = get_addresses_container(rec)
    addresses = ensure_list(addresses_container.get("address_name"))

    matching_addr_numbers: Set[str] = set()
    matching_author_keys: Set[str] = set()

    matching_author_names: Set[str] = set()
    matching_full_addresses_by_no: Dict[str, str] = {}
    matching_suborganizations: Set[str] = set()
    author_address_matches: List[str] = []

    for addr in addresses:

        if not isinstance(addr, dict):
            continue

        address_spec = addr.get("address_spec")
        if not isinstance(address_spec, dict):
            continue

        addr_no = address_spec.get("addr_no")

        # Fallback for alternate record shapes
        if addr_no is None:
            addr_no = addr.get("addr_no")

        if addr_no is None:
            continue

        addr_no = str(addr_no)
        full_address = str(address_spec.get("full_address") or "")

        org_block = address_spec.get("organizations")
        if not isinstance(org_block, dict):
            continue

        orgs = ensure_list(org_block.get("organization"))

        if not org_matches_target(orgs, target_norm):
            continue

        matching_addr_numbers.add(addr_no)

        if full_address:
            matching_full_addresses_by_no[addr_no] = full_address

        for suborg in extract_suborganizations(address_spec):
            matching_suborganizations.add(suborg)

        names_block = addr.get("names")
        if not isinstance(names_block, dict):
            continue

        names = ensure_list(names_block.get("name"))

        for name_obj in names:

            if not isinstance(name_obj, dict):
                continue

            # Usually these are authors, but keep the check defensive.
            role = normalize(name_obj.get("role"))
            if role and role != "author":
                continue

            author_key = get_author_key(name_obj)
            author_display = get_author_display(name_obj)

            if not author_key:
                continue

            matching_author_keys.add(author_key)

            if author_display:
                matching_author_names.add(author_display)

            if author_display and full_address:
                author_address_matches.append(
                    f"{author_display} => [{addr_no}] {full_address}"
                )
            elif author_display:
                author_address_matches.append(
                    f"{author_display} => [{addr_no}]"
                )

    sorted_addr_numbers = sorted(matching_addr_numbers, key=sort_addr_no)

    matching_full_addresses = [
        f"[{addr_no}] {matching_full_addresses_by_no.get(addr_no, '')}"
        for addr_no in sorted_addr_numbers
        if matching_full_addresses_by_no.get(addr_no, "")
    ]

    return {
        "total_authors": total_authors,
        "matching_authors_count": len(matching_author_keys),
        "distinct_matching_addresses_count": len(matching_addr_numbers),
        "has_2plus_matching_authors": (
            "YES" if len(matching_author_keys) >= 2 else "NO"
        ),
        "matching_author_names": safe_join(sorted(matching_author_names)),
        "matching_full_addresses": safe_join(matching_full_addresses),
        "known_suborganizations": safe_join(sorted(matching_suborganizations)),
        "author_address_matches": safe_join(sorted(set(author_address_matches))),
    }


def build_output_row(
    rec: Dict[str, Any],
    affiliation: str
) -> Tuple[List[Any], bool]:
    """
    Analyze one record and return:
    - output row
    - whether this record had at least one matching author
    """

    uid = extract_uid(rec)
    title = extract_item_title(rec)
    wos_platform_cites = extract_wos_platform_cites(rec)
    core_collection_cites = extract_core_collection_cites(rec)

    try:
        result = analyze_affiliation(rec, affiliation)

        internal_only_collab = is_internal_only_collab(
            result["total_authors"],
            result["matching_authors_count"]
        )

        row = [
            uid,
            title,
            wos_platform_cites,
            core_collection_cites,
            result["total_authors"],
            result["matching_authors_count"],
            result["distinct_matching_addresses_count"],
            result["has_2plus_matching_authors"],
            internal_only_collab,
            result["matching_author_names"],
            result["matching_full_addresses"],
            result["known_suborganizations"],
            result["author_address_matches"],
            ""
        ]

        had_matching_author = result["matching_authors_count"] > 0

        return row, had_matching_author

    except Exception as e:
        row = [
            uid,
            title,
            wos_platform_cites,
            core_collection_cites,
            "N/A",
            "N/A",
            "N/A",
            "ERROR",
            "NO",
            "",
            "",
            "",
            "",
            str(e)
        ]

        return row, False


def build_retrieval_error_row(
    first_record: int,
    attempted_count: int,
    error_message: str
) -> List[Any]:
    """
    Build an output row when a record or page cannot be retrieved.

    This allows the script to keep going and preserve a note in the workbook.
    """

    return [
        f"RETRIEVAL_ERROR_FIRST_RECORD_{first_record}",
        "",
        "",
        "",
        "N/A",
        "N/A",
        "N/A",
        "ERROR",
        "NO",
        "",
        "",
        "",
        "",
        (
            f"Could not retrieve firstRecord={first_record} even at "
            f"count={attempted_count}. {error_message}"
        )
    ]


# Excel output

HEADERS = [
    "UT",
    "Title",
    "WOS Platform Cites",
    "Core Collection Cites",
    "Total Authors",
    "Matching Authors",
    "Distinct Matching Addresses",
    ">=2 Matching Authors",
    "Internal Only Collab",
    "Matching Author Names",
    "Matching Full Addresses",
    "Known Suborganizations",
    "Author Address Matches",
    "Error"
]


def initialize_excel_output(
    output_file: Path,
    run_datetime: str,
    usr_query: str,
    affiliation: str,
    total_records: int
) -> Tuple[Workbook, Any, int]:
    """
    Create workbook, add metadata/header rows, apply formatting, and save once.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Affiliation Analysis"

    metadata_rows = [
        ["Affiliation Collaboration Analysis"],
        ["Run date/time", run_datetime],
        ["Query", usr_query],
        ["Target affiliation", affiliation],
        ["Records retrieved", total_records],
        []
    ]

    for row in metadata_rows:
        ws.append(row)

    ws.append(HEADERS)
    header_row_num = ws.max_row

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    header_font = Font(bold=True)

    ws["A1"].font = title_font
    ws["A1"].fill = title_fill

    for cell in ws[header_row_num]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Freeze above data rows and turn on filters
    ws.freeze_panes = f"A{header_row_num + 1}"
    ws.auto_filter.ref = f"A{header_row_num}:{get_column_letter(len(HEADERS))}{ws.max_row}"

    # Column widths
    column_widths = {
        "A": 22,   # UT
        "B": 55,   # Title
        "C": 16,   # WOS Platform Cites
        "D": 18,   # Core Collection Cites
        "E": 14,   # Total Authors
        "F": 16,   # Matching Authors
        "G": 24,   # Distinct Matching Addresses
        "H": 20,   # >=2 Matching Authors
        "I": 20,   # Internal Only Collab
        "J": 35,   # Matching Author Names
        "K": 60,   # Matching Full Addresses
        "L": 35,   # Known Suborganizations
        "M": 75,   # Author Address Matches
        "N": 30,   # Error
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Make metadata labels bold
    for row_num in range(2, 6):
        ws[f"A{row_num}"].font = Font(bold=True)

    # Reasonable row heights
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[header_row_num].height = 32

    wb.save(output_file)

    return wb, ws, header_row_num


def append_data_row(ws: Any, row: List[Any], header_row_num: int) -> None:
    """
    Append one output row and apply row-level alignment.
    """

    ws.append(row)
    row_num = ws.max_row

    wrap_cols = {"B", "J", "K", "L", "M", "N"}
    center_cols = {"C", "D", "E", "F", "G", "H", "I"}

    for cell in ws[row_num]:
        col_letter = get_column_letter(cell.column)

        if col_letter in wrap_cols:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        elif col_letter in center_cols:
            cell.alignment = Alignment(horizontal="center", vertical="top")
        else:
            cell.alignment = Alignment(vertical="top")

    # Keep filter range current as rows are added
    ws.auto_filter.ref = f"A{header_row_num}:{get_column_letter(len(HEADERS))}{ws.max_row}"


# Command-line entry point

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Analyze affiliation collaboration using Web of Science "
            "Expanded API records."
        )
    )

    parser.add_argument(
        "-q",
        "--query",
        default=DEFAULT_QUERY,
        help=f"WoS query. Default: {DEFAULT_QUERY}"
    )

    parser.add_argument(
        "-a",
        "--affiliation",
        default=DEFAULT_AFFILIATION,
        help=f"Target affiliation. Default: {DEFAULT_AFFILIATION}"
    )

    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help=(
            "Output Excel file. If omitted, a timestamped filename is "
            "created automatically."
        )
    )

    parser.add_argument(
        "--count",
        type=int,
        default=COUNT,
        help=f"Records per API page. Default: {COUNT}"
    )

    parser.add_argument(
        "--yes",
        action="store_true",
        help="Continue past guardrail warnings without prompting."
    )

    parser.add_argument(
        "--skip-affiliation-check",
        action="store_true",
        help="Skip the OG=(affiliation) preflight check."
    )

    parser.add_argument(
        "--skip-main-query-count-check",
        action="store_true",
        help="Skip the main query size preflight check."
    )

    parser.add_argument(
        "--affiliation-warning-threshold",
        type=int,
        default=AFFILIATION_WARNING_THRESHOLD,
        help=(
            "Warn if OG=(affiliation) returns fewer than this many records. "
            f"Default: {AFFILIATION_WARNING_THRESHOLD}"
        )
    )

    parser.add_argument(
        "--main-query-warning-threshold",
        type=int,
        default=MAIN_QUERY_WARNING_THRESHOLD,
        help=(
            "Warn if the main query returns this many records or more. "
            f"Default: {MAIN_QUERY_WARNING_THRESHOLD}"
        )
    )

    args = parser.parse_args()

    load_dotenv()
    apikey = os.getenv("EXPANDED_APIKEY")
    if not apikey:
        raise RuntimeError("Missing EXPANDED_APIKEY in .env")

    usr_query = args.query
    affiliation = args.affiliation

    if not usr_query.strip():
        raise RuntimeError("Query is blank. Please provide a valid WoS query.")

    if not affiliation.strip():
        raise RuntimeError("Affiliation is blank. Please provide a valid target affiliation.")

    if args.output:
        output_file = Path(args.output)
    else:
        output_file = make_default_output_filename()

    # Add the Excel extension when omitted.
    if output_file.suffix.lower() != ".xlsx":
        output_file = output_file.with_suffix(".xlsx")

    run_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    print("----- INPUTS -----")
    print(f"Run date/time: {run_datetime}")
    print(f"Query: {usr_query}")
    print(f"Affiliation: {affiliation}")
    print(f"Count: {args.count}")
    print(f"Output file: {output_file}")
    print("------------------")

    # Validate the target affiliation before the main retrieval.
    if not args.skip_affiliation_check:
        try:
            continue_run = run_affiliation_preflight_check(
                apikey=apikey,
                affiliation=affiliation,
                threshold=args.affiliation_warning_threshold,
                auto_yes=args.yes
            )
        except InvalidWoSQueryError as e:
            print(e)
            print("No output file was created.")
            return
        except WoSAuthenticationError as e:
            print(e)
            print("No output file was created.")
            return
        except Exception as e:
            print()
            print("Affiliation preflight check failed.")
            print(str(e))
            print("No output file was created.")
            return

        if not continue_run:
            print("Run cancelled. No output file was created.")
            return
    else:
        print("Skipping affiliation preflight check because --skip-affiliation-check was used.")

    # Warn before retrieving a very large result set.
    if not args.skip_main_query_count_check:
        try:
            continue_run = run_main_query_count_preflight_check(
                apikey=apikey,
                usr_query=usr_query,
                threshold=args.main_query_warning_threshold,
                auto_yes=args.yes
            )
        except InvalidWoSQueryError as e:
            print(e)
            print("No output file was created.")
            return
        except WoSAuthenticationError as e:
            print(e)
            print("No output file was created.")
            return
        except Exception as e:
            print()
            print("Main query size preflight check failed.")
            print(str(e))
            print("No output file was created.")
            return

        if not continue_run:
            print("Run cancelled. No output file was created.")
            return
    else:
        print("Skipping main query size preflight check because --skip-main-query-count-check was used.")

    params = {
        "databaseId": "WOS",
        "usrQuery": usr_query
    }

    print("Retrieving and processing records page by page...")

    first_response, first_page_count, first_page_error = get_wos_page_with_auto_shrink(
        apikey=apikey,
        params=params,
        first_record=1,
        preferred_count=args.count
    )

    if first_response is None:
        print()
        print("Could not retrieve the first page, even after shrinking page size.")
        print(first_page_error)
        print("No output file was created.")
        return

    total_records = extract_records_found(first_response)
    print(f"Total records found: {total_records}")

    if total_records == 0:
        print("No results were retrieved for this query. No output file was created.")
        return

    if total_records > 100000:
        print(f"Error: Query returned {total_records} results, max allowed is 100000.")
        print("No output file was created.")
        return

    wb, ws, header_row_num = initialize_excel_output(
        output_file=output_file,
        run_datetime=run_datetime,
        usr_query=usr_query,
        affiliation=affiliation,
        total_records=total_records
    )

    any_matching_authors = False
    processed_count = 0

    first_record = 1

    current_count = args.count
    successful_pages_at_current_count = 0

    while first_record <= total_records:

        if first_record == 1:
            response = first_response
            page_count_used = first_page_count
        else:
            response, page_count_used, page_error = get_wos_page_with_auto_shrink(
                apikey=apikey,
                params=params,
                first_record=first_record,
                preferred_count=current_count
            )

            if response is None:
                print()
                print(f"Could not retrieve firstRecord={first_record} even at count=1.")
                print("Logging retrieval error row and continuing.")

                error_row = build_retrieval_error_row(
                    first_record=first_record,
                    attempted_count=1,
                    error_message=page_error or "Unknown retrieval error"
                )

                append_data_row(ws, error_row, header_row_num)
                processed_count += 1

                wb.save(output_file)
                print(f"Saved {processed_count} of {total_records} records to {output_file}")

                # Skip the single record that could not be retrieved.
                first_record += 1
                continue

        recs = extract_records_from_response(response)

        if not recs:
            print(f"No 'REC' in response for firstRecord = {first_record}")

            error_row = build_retrieval_error_row(
                first_record=first_record,
                attempted_count=page_count_used,
                error_message="No REC records found in API response."
            )

            append_data_row(ws, error_row, header_row_num)
            processed_count += page_count_used

        else:
            for rec in recs:
                row, had_matching_author = build_output_row(rec, affiliation)
                append_data_row(ws, row, header_row_num)

                processed_count += 1

                if had_matching_author:
                    any_matching_authors = True

        # Save after every successful page.
        wb.save(output_file)
        print(f"Saved {processed_count} of {total_records} records to {output_file}")

        if page_count_used < current_count:
            current_count = page_count_used
            successful_pages_at_current_count = 0
            print(f"Adaptive page size reduced to {current_count}.")
        else:
            successful_pages_at_current_count += 1

        if successful_pages_at_current_count >= ADAPTIVE_STEP_UP_AFTER_PAGES:
            next_count = get_next_larger_page_size(current_count, args.count)

            if next_count > current_count:
                print(f"Adaptive page size stepping up from {current_count} to {next_count}.")
                current_count = next_count

            successful_pages_at_current_count = 0
        # Advance by the page size that actually succeeded.
        first_record += page_count_used

    if not any_matching_authors:
        print()
        print("----- WARNING -----")
        print(
            "Records were retrieved, but no matching authors were found for the "
            "target affiliation. The output file was still created because "
            "the retrieved records may be useful for troubleshooting."
        )

    print(f"Done. Final output file: {output_file}")


if __name__ == "__main__":
    main()